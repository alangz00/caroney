import streamlit as st
import pandas as pd
from io import BytesIO
import datetime
import json
import gspread
from oauth2client.service_account import ServiceAccountCredentials

# Formato Excel (openpyxl)
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, Border, Side

# =========================
# Config general de la app
# =========================
st.set_page_config(page_title="Caroney", page_icon="🐎", layout="centered")
st.title("💸 - Caroney - Feliz Espuki Sison")
st.markdown("Registra tus ingresos y egresos de forma compacta y bonita. ¡Hecho con cariño!")
# Botón de refrescar datos
if st.button("🔄 Actualizar datos"):
    if "records" in st.session_state:
        del st.session_state["records"]
    st.rerun()



# =====================================
# 🎃 Mensaje de Octubre con GIF aleatorio
# =====================================
import random


st.markdown("---")
st.subheader("🎃 Especial de Octubre")

gifs = [
    # 🐱 Gatitos
    "https://media.giphy.com/media/JIX9t2j0ZTN9S/giphy.gif",
    "https://media.giphy.com/media/v6aOjy0Qo1fIA/giphy.gif",
    "https://media.giphy.com/media/mlvseq9yvZhba/giphy.gif",
    # 🐶 Perritos
    "https://media.giphy.com/media/26FPqut4b5p7K6kju/giphy.gif",
    "https://media.giphy.com/media/26tPplGWjN0xLybiU/giphy.gif",
    # 🐴 Caballos
    "https://giphy.com/gifs/funny-dance-weird-xqBpxjk7CXLtm",
    "https://media.giphy.com/media/3orieYlY4JG1N2VR0Q/giphy.gif",
    # 🌸 Otros bonitos
    "https://media.giphy.com/media/MDJ9IbxxvDUQM/giphy.gif",
    "https://media.giphy.com/media/l0MYEqEzwMWFCg8rm/giphy.gif"
]

if st.button("Click aquí para una sorpresa 👻"):
    gif_url = random.choice(gifs)
    st.markdown("¡Feliz Octubre! 🍂🍁")
    st.image(gif_url, caption="¡Sorpresa! 🎃")

    
    
# ==========================
# ==========================


# 🔐 Conectar con Google Sheets
scope = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]
service_account_info = json.loads(st.secrets["GOOGLE_CREDENTIALS"])
creds = ServiceAccountCredentials.from_json_keyfile_dict(service_account_info, scope)
client = gspread.authorize(creds)
sheet = client.open("carodb").sheet1  # Asegúrate que así se llame tu hoja

# =========================
# Carga de datos
# =========================
if "records" not in st.session_state:
    sheet_data = sheet.get_all_records()  # sin filas vacías y con encabezados de la fila 1
    st.session_state.records = sheet_data

# Construir DF base
df = pd.DataFrame(st.session_state.records) if st.session_state.records else pd.DataFrame(
    columns=["Fecha", "Monto", "Tipo", "Categoría", "Descripción"]
)
if not df.empty:
    # Tipos y columnas
    df["Fecha"] = pd.to_datetime(df["Fecha"])
    df["Monto"] = pd.to_numeric(df["Monto"], errors="coerce")

    # Mapear filas de Google Sheets (2..N+1) para poder EDITAR/ELIMINAR en origen
    # get_all_records ignora filas vacías, pero mantiene el orden; asumimos hoja limpia sin huecos intermedios.
    df["GSRow"] = range(2, 2 + len(df))  # 1 es encabezado

# =========================
# Formulario de entrada
# =========================
with st.form("entry_form"):
    col1, col2 = st.columns(2)
    with col1:
        date = st.date_input("Fecha")
        amount = st.number_input("Monto", min_value=0.0, step=0.01)
    with col2:
        type_ = st.selectbox("Tipo", ["Ingreso", "Egreso"])
        category = st.text_input("Categoría (ej. comida, renta)")

    description = st.text_input("Descripción")
    submitted = st.form_submit_button("Agregar")

    if submitted:
        # Construir la fila limpiamente
        signed_amount = float(amount if type_ == "Ingreso" else -amount)
        row = [
            str(date),
            signed_amount,
            str(type_).strip(),
            str(category).strip() if category else "Sin categoría",
            str(description).strip() if description else "",
        ]

        # Guardar en Google Sheets (primero origen)
        sheet.append_row(row)

        # Actualizar estado local (agregamos GSRow = última fila)
        new_gsrow = 2 + len(df) + 1 if not df.empty else 2
        new_entry = {
            "Fecha": str(date),
            "Monto": signed_amount,
            "Tipo": type_,
            "Categoría": row[3],
            "Descripción": row[4],
        }
        if df.empty:
            df = pd.DataFrame([new_entry])
            df["Fecha"] = pd.to_datetime(df["Fecha"])
            df["Monto"] = pd.to_numeric(df["Monto"], errors="coerce")
            df["GSRow"] = [new_gsrow]
        else:
            add_row = pd.DataFrame([new_entry])
            df = pd.concat([df, add_row], ignore_index=True)
            df.loc[df.index[-1], "Fecha"] = pd.to_datetime(df.loc[df.index[-1], "Fecha"])
            df.loc[df.index[-1], "GSRow"] = new_gsrow

        # Actualizar session_state.records desde df (sin GSRow)
        st.session_state.records = df.drop(columns=["GSRow"]).to_dict(orient="records")
        st.success("Movimiento agregado ✅")

# =========================
# Mostrar datos
# =========================
if df.empty:
    st.info("Aún no has registrado nada.")
    st.stop()

# =========================
# 📅 Balance del MES actual
# =========================
hoy = datetime.date.today()
primer_dia_mes = hoy.replace(day=1)
filtro_mes = (df["Fecha"].dt.date >= primer_dia_mes) & (df["Fecha"].dt.date <= hoy)
df_mes = df[filtro_mes].copy()

st.subheader("📆 Movimientos del mes (hasta hoy)")

if df_mes.empty:
    st.info("Aún no hay movimientos este mes.")
else:
    st.dataframe(df_mes.drop(columns=["GSRow"]), use_container_width=True)

    ingresos_mes = df_mes[df_mes["Tipo"] == "Ingreso"]["Monto"].sum()
    egresos_mes = -df_mes[df_mes["Tipo"] == "Egreso"]["Monto"].sum()
    balance_mes = df_mes["Monto"].sum()

    st.markdown(f"**Ingresos del mes:** ${ingresos_mes:.2f}")
    st.markdown(f"**Egresos del mes:** ${egresos_mes:.2f}")
    st.markdown(f"**Balance del mes:** ${balance_mes:.2f}")

    # Excel del mes (bonito + resumen) con nombre por mes
    meses_es = ["", "enero", "febrero", "marzo", "abril", "mayo", "junio",
                "julio", "agosto", "septiembre", "octubre", "noviembre", "diciembre"]
    mes_nombre = f"{meses_es[hoy.month]} {hoy.year}"

    df_mes_export = df_mes.drop(columns=["GSRow"]).copy()
    df_mes_export["Fecha"] = pd.to_datetime(df_mes_export["Fecha"]).dt.date

    resumen_mes = pd.DataFrame(
        [
            {"Fecha": "MES", "Monto": ingresos_mes, "Tipo": "Ingreso", "Descripción": "Ingresos del mes"},
            {"Fecha": "MES", "Monto": egresos_mes, "Tipo": "Egreso", "Descripción": "Egresos del mes"},
            {"Fecha": "MES", "Monto": balance_mes, "Descripción": "Balance neto del mes"},
        ]
    )

    df_mes_export = pd.concat([df_mes_export, pd.DataFrame([{}]), resumen_mes], ignore_index=True)

    wb_mes = Workbook()
    ws_mes = wb_mes.active
    ws_mes.title = "Caroney Mes"

    for r in dataframe_to_rows(df_mes_export, index=False, header=True):
        ws_mes.append(r)

    # Encabezados en negritas y centrados
    for cell in ws_mes[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # Auto-anchos
    for col in ws_mes.columns:
        max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws_mes.column_dimensions[col[0].column_letter].width = max_len + 2

    # Bordes delgados
    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"),
                         top=Side(style="thin"), bottom=Side(style="thin"))
    for row_ in ws_mes.iter_rows(min_row=1, max_row=ws_mes.max_row, min_col=1, max_col=ws_mes.max_column):
        for cell in row_:
            if cell.value is not None:
                cell.border = thin_border

    towrite_mes = BytesIO()
    wb_mes.save(towrite_mes)
    towrite_mes.seek(0)
    st.download_button("📥 Descargar Excel del mes", towrite_mes, f"caroney_mes_{mes_nombre}.xlsx")

# ===================================
# 📆 Filtro por fechas + Excel bonito
# ===================================
if "mostrar_filtro" not in st.session_state:
    st.session_state.mostrar_filtro = False

if st.button("📆 Filtrar por fechas"):
    st.session_state.mostrar_filtro = not st.session_state.mostrar_filtro

if st.session_state.mostrar_filtro:
    min_date = df["Fecha"].min().date()
    max_date = df["Fecha"].max().date()

    start_date, end_date = st.date_input(
        "Selecciona el rango:",
        value=(min_date, max_date),
        min_value=min_date,
        max_value=max_date,
    )

    filtro = (df["Fecha"].dt.date >= start_date) & (df["Fecha"].dt.date <= end_date)
    df_filtro = df[filtro].copy()

    st.subheader("📆 Movimientos filtrados")
    if df_filtro.empty:
        st.info("No hay movimientos en el rango seleccionado.")
    else:
        st.dataframe(df_filtro.drop(columns=["GSRow"]), use_container_width=True)

        ingresos_f = df_filtro[df_filtro["Tipo"] == "Ingreso"]["Monto"].sum()
        egresos_f = -df_filtro[df_filtro["Tipo"] == "Egreso"]["Monto"].sum()
        balance_f = df_filtro["Monto"].sum()

        st.markdown(f"**Ingresos filtrados:** ${ingresos_f:.2f}")
        st.markdown(f"**Egresos filtrados:** ${egresos_f:.2f}")
        st.markdown(f"**Balance filtrado:** ${balance_f:.2f}")

        # ---------- Excel filtrado BONITO (formato + resumen) ----------
        df_filtro_export = df_filtro.drop(columns=["GSRow"]).copy()
        df_filtro_export["Fecha"] = pd.to_datetime(df_filtro_export["Fecha"]).dt.date

        resumen_f = pd.DataFrame([
            {"Fecha": "RANGO", "Monto": ingresos_f, "Tipo": "Ingreso", "Descripción": "Ingresos (rango)"},
            {"Fecha": "RANGO", "Monto": egresos_f, "Tipo": "Egreso", "Descripción": "Egresos (rango)"},
            {"Fecha": "RANGO", "Monto": balance_f, "Descripción": "Balance neto (rango)"}
        ])

        df_filtro_export = pd.concat([df_filtro_export, pd.DataFrame([{}]), resumen_f], ignore_index=True)

        wb_f = Workbook()
        ws_f = wb_f.active
        ws_f.title = "Caroney Rango"

        for r in dataframe_to_rows(df_filtro_export, index=False, header=True):
            ws_f.append(r)

        # Encabezados: negritas + centrado
        for cell in ws_f[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        # Auto ancho de columnas
        for col in ws_f.columns:
            max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            ws_f.column_dimensions[col[0].column_letter].width = max_len + 2

        # Bordes delgados
        thin_border = Border(left=Side(style="thin"), right=Side(style="thin"),
                             top=Side(style="thin"), bottom=Side(style="thin"))
        for row_ in ws_f.iter_rows(min_row=1, max_row=ws_f.max_row, min_col=1, max_col=ws_f.max_column):
            for cell in row_:
                if cell.value is not None:
                    cell.border = thin_border

        # Nombre de archivo con rango
        fname = f"caroney_filtrado_{start_date.isoformat()}_a_{end_date.isoformat()}.xlsx"

        towrite = BytesIO()
        wb_f.save(towrite)
        towrite.seek(0)
        st.download_button("📥 Descargar Excel filtrado", towrite, fname)

# =====================================
# 📖 Historial completo + Excel bonito
# =====================================
if "mostrar_historial_completo" not in st.session_state:
    st.session_state.mostrar_historial_completo = False

if st.button("📖 Ver todos los movimientos"):
    st.session_state.mostrar_historial_completo = not st.session_state.mostrar_historial_completo

if st.session_state.mostrar_historial_completo:
    st.subheader("📋 Historial completo")
    st.dataframe(df.drop(columns=["GSRow"]), use_container_width=True)

    total_ingresos = df[df["Tipo"] == "Ingreso"]["Monto"].sum()
    total_egresos = -df[df["Tipo"] == "Egreso"]["Monto"].sum()
    balance_total = df["Monto"].sum()

    st.markdown(f"**Total de ingresos:** ${total_ingresos:.2f}")
    st.markdown(f"**Total de egresos:** ${total_egresos:.2f}")
    st.markdown(f"**Balance general:** ${balance_total:.2f}")

    df_export = df.drop(columns=["GSRow"]).copy()
    df_export["Fecha"] = pd.to_datetime(df_export["Fecha"]).dt.date

    resumen = pd.DataFrame(
        [
            {"Fecha": "TOTAL", "Monto": total_ingresos, "Tipo": "Ingreso", "Descripción": "Ingresos totales"},
            {"Fecha": "TOTAL", "Monto": total_egresos, "Tipo": "Egreso", "Descripción": "Egresos totales"},
            {"Fecha": "TOTAL", "Monto": balance_total, "Descripción": "Balance neto"},
        ]
    )

    df_export = pd.concat([df_export, pd.DataFrame([{}]), resumen], ignore_index=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Caroney"

    for r in dataframe_to_rows(df_export, index=False, header=True):
        ws.append(r)

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    for col in ws.columns:
        max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_len + 2

    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"),
                         top=Side(style="thin"), bottom=Side(style="thin"))
    for row_ in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row_:
            if cell.value is not None:
                cell.border = thin_border

    towrite_full = BytesIO()
    wb.save(towrite_full)
    towrite_full.seek(0)
    st.download_button("📥 Descargar Excel completo", towrite_full, "caroney_completo.xlsx")

# =====================================
# ✏️ Editar / 🗑️ Eliminar movimientos
# =====================================
st.markdown("---")
st.subheader("✏️ Editar o 🗑️ Eliminar un movimiento")

if df.empty:
    st.info("No hay movimientos para editar o eliminar.")
else:
    # Vista amigable para seleccionar
    df_view = df.copy()
    df_view["Fecha"] = df_view["Fecha"].dt.strftime("%Y-%m-%d")
    options = [
        f"{i}. {row.Fecha} | {row.Tipo} | ${row.Monto:.2f} | {row['Categoría']} | {row['Descripción']}"
        for i, row in df_view.iterrows()
    ]
    selected = st.selectbox("Elige un movimiento:", options, index=None, placeholder="Selecciona una fila…")

    if selected is not None:
        idx = int(selected.split(".")[0])  # índice en df
        row = df.iloc[idx]
        gsrow = int(row["GSRow"])

        # Valores actuales
        cur_date = pd.to_datetime(row["Fecha"]).date()
        cur_type = row["Tipo"]
        cur_amount_abs = float(abs(row["Monto"]))
        cur_cat = row["Categoría"]
        cur_desc = row["Descripción"]

        with st.form("edit_form"):
            c1, c2 = st.columns(2)
            with c1:
                new_date = st.date_input("Fecha", value=cur_date)
                new_amount_abs = st.number_input("Monto (positivo)", min_value=0.0, step=0.01, value=cur_amount_abs)
            with c2:
                new_type = st.selectbox("Tipo", ["Ingreso", "Egreso"], index=0 if cur_type == "Ingreso" else 1)
                new_cat = st.text_input("Categoría", value=cur_cat)

            new_desc = st.text_input("Descripción", value=cur_desc)

            col_a, col_b, col_c = st.columns([1, 1, 2])
            save = col_a.form_submit_button("💾 Guardar cambios")
            delete = col_b.form_submit_button("🗑️ Eliminar")

        if save:
            # Firmar el monto según tipo
            signed_amount = float(new_amount_abs if new_type == "Ingreso" else -new_amount_abs)

            # Actualizar en Google Sheets (A..E)
            sheet.update(f"A{gsrow}:E{gsrow}", [[
                str(new_date),
                signed_amount,
                new_type,
                new_cat if new_cat else "Sin categoría",
                new_desc if new_desc else ""
            ]])

            st.success("Movimiento actualizado ✅")
            st.rerun()

        if delete:
            # Eliminar en Google Sheets
            sheet.delete_rows(gsrow)
            st.success("Movimiento eliminado ✅")
            st.rerun()

